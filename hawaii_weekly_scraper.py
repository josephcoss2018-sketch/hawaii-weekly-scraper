#!/usr/bin/env python3
"""
Hawaii Campaign Finance – 2026 Candidate Report Scraper
Source: https://olvr.hawaii.gov/Controls/CandidateFiling.aspx?elid=94
Runs weekly via GitHub Actions; auto-exits after STOP_DATE.
"""

import datetime
import time
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ── config ────────────────────────────────────────────────────────────────────────────────
BASE_URL  = "https://olvr.hawaii.gov/Controls/CandidateFiling.aspx?elid=94"
STOP_DATE = datetime.date(2026, 9, 1)
OUT_DIR   = Path("hawaii_reports")
COLS      = ["Contest", "Party", "Ballot Name", "Legal Name",
             "Mailing Address", "Phone", "Email", "Website",
             "Issued", "Filed", "Status"]


def main():
    today = datetime.date.today()
    if today >= STOP_DATE:
        print(f"STOP_DATE {STOP_DATE} reached – exiting.")
        return

    OUT_DIR.mkdir(exist_ok=True)
    print(f"[{today}] Starting Hawaii 2026 Candidate Report scrape…")

    records = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(BASE_URL, wait_until="networkidle", timeout=60_000)
        time.sleep(2)

        page_num = 1
        while True:
            print(f"  Scraping page {page_num}…", end=" ", flush=True)
            rows = extract_page_rows(page)
            records.extend(rows)
            print(f"{len(rows)} rows (total {len(records)})")

            if not click_next_page(page, page_num):
                print(f"  → Finished at page {page_num}.")
                break
            page_num += 1

        browser.close()

    print(f"\nTotal records collected: {len(records)}")
    out_path = OUT_DIR / f"hawaii_candidates_{today.strftime('%Y%m%d')}.xlsx"
    write_excel(records, out_path, today)
    print(f"Excel saved: {out_path}")


def get_current_page_number(page) -> str:
    """Return the text of the current-page indicator in the Telerik pager."""
    try:
        el = page.query_selector("a.rgCurrentPage span")
        if el:
            return el.inner_text().strip()
        el = page.query_selector(".rgCurrentPage span")
        return el.inner_text().strip() if el else ""
    except Exception:
        return ""


def extract_page_rows(page) -> list:
    """Pull all 11-column rows from the current grid page."""
    cells = page.query_selector_all("table.rgMasterTable tbody tr td")
    if not cells:
        return []

    values = []
    for cell in cells:
        link = cell.query_selector("a[href^='mailto:']")
        if link:
            href = link.get_attribute("href") or ""
            txt = href.replace("mailto:", "").strip()
        else:
            txt = cell.inner_text().strip()
        values.append(txt)

    rows = []
    n = len(COLS)
    for i in range(0, len(values) - n + 1, n):
        chunk = values[i : i + n]
        if len(chunk) == n:
            rows.append(dict(zip(COLS, chunk)))
    return rows


def click_next_page(page, current_page_num: int) -> bool:
    """
    Click the Telerik 'Next Page' button (input.rgPageNext) and wait for the
    grid to refresh.  Returns False when already on the last page.
    """
    btn = page.query_selector("input.rgPageNext")
    if not btn:
        return False

    if page.evaluate("el => el.disabled", btn):
        return False

    page_before = get_current_page_number(page) or str(current_page_num)
    btn.click()

    try:
        page.wait_for_function(
            f"""() => {{
                const el = document.querySelector('a.rgCurrentPage span');
                if (!el) return false;
                const txt = el.innerText.trim();
                return txt !== '' && txt !== {repr(page_before)};
            }}""",
            timeout=20_000,
        )
        return True
    except PlaywrightTimeoutError:
        return False


def write_excel(records: list, out_path: Path, run_date: datetime.date):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Hawaii 2026 Candidates"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    alt_fill  = PatternFill("solid", fgColor="D9E1F2")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")

    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for r_idx, rec in enumerate(records, start=2):
        for c_idx, col in enumerate(COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=rec.get(col, ""))
            cell.alignment = Alignment(vertical="top")
            cell.fill = alt_fill if r_idx % 2 == 0 else norm_fill

    widths = {"Contest": 35, "Party": 14, "Ballot Name": 28, "Legal Name": 28,
              "Mailing Address": 42, "Phone": 16, "Email": 36, "Website": 36,
              "Issued": 14, "Filed": 14, "Status": 12}
    for c_idx, col in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 20)

    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12

    ws2["A1"] = "Hawaii 2026 Candidate Report"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws2.merge_cells("A1:B1")

    ws2.append([])
    ws2.append(["Run Date",      str(run_date)])
    ws2.append(["Total Records", len(records)])
    ws2.append(["Source URL",    BASE_URL])
    ws2.append([])

    sec_font = Font(bold=True, size=11, color="1F4E79")

    def section(label, counter_key):
        ws2.append([label, "Count"])
        r = ws2.max_row
        ws2.cell(row=r, column=1).font = sec_font
        ws2.cell(row=r, column=2).font = sec_font
        for name, cnt in sorted(Counter(rec.get(counter_key, "") for rec in records).items(),
                                key=lambda x: -x[1]):
            ws2.append([name, cnt])
        ws2.append([])

    section("Party Breakdown",   "Party")
    section("Status Breakdown",  "Status")
    section("Contest Breakdown", "Contest")

    wb.save(out_path)


if __name__ == "__main__":
    main()
